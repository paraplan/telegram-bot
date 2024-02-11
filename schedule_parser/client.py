import asyncio
from io import BytesIO
from typing import Final

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from daemon.__main__ import get_schedules

work_book: Final[Workbook] = load_workbook(BytesIO((asyncio.run(get_schedules())[0])))
sheets: list[str] = work_book.sheetnames
work_sheet: Final[Worksheet] = work_book[sheets[0]]
